# analyzer/core.py
import os, math, json, glob
import cv2
import numpy as np
import pandas as pd
from PIL import Image
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ====== 원본 코드에서 가져온(또는 축약한) 핵심 유틸 ======
def safe_fps(cap):
    fps = cap.get(cv2.CAP_PROP_FPS)
    return fps if fps and fps > 0 else 30.0

def video_meta(path):
    cap = cv2.VideoCapture(path)
    if not cap.isOpened():
        return {"opened": False, "fps": 0.0, "frames": 0, "secs": 0.0}
    fps = safe_fps(cap)
    total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    secs = (total / fps) if fps > 0 else 0
    cap.release()
    return {"opened": True, "fps": fps, "frames": total, "secs": secs}

def tenengrad_energy(gray, mask=None):
    sx = cv2.Sobel(gray, cv2.CV_64F, 1, 0, ksize=3)
    sy = cv2.Sobel(gray, cv2.CV_64F, 0, 1, ksize=3)
    energy = sx*sx + sy*sy
    if mask is not None:
        vals = energy[mask]
        return float(np.mean(vals)) if vals.size > 0 else 0.0
    return float(np.mean(energy))

def highfreq_noise(gray, mask=None):
    blur = cv2.medianBlur(gray, 5)
    diff = gray.astype(np.float32) - blur.astype(np.float32)
    if mask is not None:
        diff = diff[mask]
        return float(np.std(diff)) if diff.size > 0 else 0.0
    return float(np.std(diff))

def spatial_uniformity(gray, tile=32, mask=None):
    h, w = gray.shape
    means = []
    for y in range(0, h, tile):
        for x in range(0, w, tile):
            patch = gray[y:min(y+tile, h), x:min(x+tile, w)]
            if mask is not None:
                mpatch = mask[y:min(y+tile, h), x:min(x+tile, w)]
                if mpatch.size < 32 or np.count_nonzero(mpatch) < 8:
                    continue
                vals = patch[mpatch]
                means.append(np.mean(vals))
            else:
                if patch.size >= 32:
                    means.append(np.mean(patch))
    return float(np.std(means)) if means else 0.0

def clip_ratio(gray, lo=3, hi=252, mask=None):
    g = gray[mask] if (mask is not None) else gray
    if g.size == 0:
        return 0.0
    return float(np.mean((g <= lo) | (g >= hi)))

def score_mono(value, p10, p90, direction="low_is_better"):
    if p90 - p10 < 1e-6:
        s = 1.0
    else:
        s = (value - p10) / (p90 - p10)
        s = max(0.0, min(1.0, s))
    if direction == "low_is_better":
        s = 1.0 - s
    return 100.0 * s

def score_twosided(value, target, sigma):
    z = (value - target) / (sigma + 1e-6)
    s = math.exp(-0.5 * (z*z))
    return 100.0 * s

# ----- 간소화: 레퍼런스 없이도 동작하도록 중앙분위 기반 자동 스케일 -----
def auto_reference_from_metrics(metrics_df):
    """
    업로드한 한 개의 비디오만으로도 동작하도록,
    프레임들의 분위수 기반 구간을 자동 설정합니다.
    """
    feats = {}
    for col in ["grad_energy","noise_hf","uniformity","brightness","flicker","roi_violation","top_badness","bottom_badness"]:
        if col not in metrics_df: 
            continue
        arr = metrics_df[col].to_numpy()
        p10 = float(np.percentile(arr, 10))
        p90 = float(np.percentile(arr, 90))
        feats[col] = {"p10":p10, "p90":p90}
    # 밝기(양측) 타깃/시그마 추정
    if "brightness" in metrics_df:
        b = metrics_df["brightness"].to_numpy()
        feats["brightness"]["target"] = float(np.median(b))
        # IQR → sigma 근사
        q25, q75 = np.percentile(b, [25, 75])
        sigma = float(max((q75-q25)/1.349, 5.0))
        feats["brightness"]["sigma"] = sigma
    return feats

def extract_metrics_per_frame(video_path, start_sec=3, max_frames=30, mask_bool=None, zone_size=0.30):
    meta = video_meta(video_path)
    if not meta["opened"]:
        raise RuntimeError(f"영상 열기 실패: {video_path}")
    cap = cv2.VideoCapture(video_path)
    fps = meta["fps"]
    start_frame = int(fps * start_sec)
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    if total_frames > 0 and start_frame >= total_frames:
        start_frame = 0

    rows = []
    idx = 0
    k = 0
    prev_brightness = None
    prev_gray = None

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        if idx >= start_frame and k < max_frames:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            if mask_bool is not None and np.count_nonzero(mask_bool) > 0:
                gvals = gray[mask_bool]
                brightness = float(np.mean(gvals))
                grad_energy = tenengrad_energy(gray, mask=mask_bool)
                noise_hf_val = highfreq_noise(gray, mask=mask_bool)
                uniformity = spatial_uniformity(gray, tile=32, mask=mask_bool)
                clip_pct = clip_ratio(gray, mask=mask_bool)
            else:
                brightness = float(np.mean(gray))
                grad_energy = tenengrad_energy(gray)
                noise_hf_val = highfreq_noise(gray)
                uniformity = spatial_uniformity(gray, tile=32, mask=None)
                clip_pct = clip_ratio(gray, mask=None)

            # flicker
            flicker = 0.0 if prev_brightness is None else float(abs(brightness - prev_brightness))
            prev_brightness = brightness

            # 간소화: ROI 위반·상/하단 품질은 placeholder (원본 세그멘테이션 필요시 이 구간 교체)
            edges = cv2.Canny(gray, 50, 150)
            roi_violation = 0.0
            top_badness = float(np.std(edges[: int(gray.shape[0]*zone_size), :]))
            bottom_badness = float(np.std(edges[int(gray.shape[0]*(1-zone_size)) :, :]))

            rows.append({
                "frame_index": idx,
                "brightness": brightness,
                "grad_energy": grad_energy,
                "noise_hf": noise_hf_val,
                "uniformity": uniformity,
                "flicker": flicker,
                "clip_pct": clip_pct,
                "roi_violation": roi_violation,
                "top_badness": top_badness,
                "bottom_badness": bottom_badness,
            })
            k += 1
        idx += 1

    cap.release()
    return pd.DataFrame(rows)

def score_frames(df, feats):
    out = []
    for _, r in df.iterrows():
        s_smooth = score_mono(r["grad_energy"], feats["grad_energy"]["p10"], feats["grad_energy"]["p90"], "low_is_better")
        s_clean  = score_mono(r["noise_hf"],   feats["noise_hf"]["p10"],   feats["noise_hf"]["p90"],   "low_is_better")
        s_uniform= score_mono(r["uniformity"], feats["uniformity"]["p10"], feats["uniformity"]["p90"], "low_is_better")
        s_stable = score_mono(r["flicker"],    feats["flicker"]["p10"],    feats["flicker"]["p90"],    "low_is_better")
        s_prem   = score_twosided(r["brightness"], feats["brightness"]["target"], feats["brightness"]["sigma"])
        if r["clip_pct"] > 0.05:
            s_prem *= max(0.0, 1.0 - (r["clip_pct"] - 0.05) * 2.0)
        s_prem = max(0.0, min(100.0, s_prem))

        # 간소화 스코어
        s_roi   = score_mono(r["roi_violation"], feats["roi_violation"]["p10"], feats["roi_violation"]["p90"], "low_is_better")
        s_top   = score_mono(r["top_badness"],   feats["top_badness"]["p10"],   feats["top_badness"]["p90"],   "low_is_better")
        s_bottom= score_mono(r["bottom_badness"],feats["bottom_badness"]["p10"],feats["bottom_badness"]["p90"],"low_is_better")

        simple_avg = float(np.mean([s_smooth, s_clean, s_uniform, s_stable, s_prem, s_roi, s_top, s_bottom]))

        out.append({
            "프레임번호": int(r["frame_index"]),
            "매끈한": round(s_smooth,1),
            "깨끗한": round(s_clean,1),
            "일정한": round(s_uniform,1),
            "안정적인": round(s_stable,1),
            "고급스러운": round(s_prem,1),
            "영역준수": round(s_roi,1),
            "상단품질": round(s_top,1),
            "하단품질": round(s_bottom,1),
            "총점(평균)": round(simple_avg,1),
        })
    return pd.DataFrame(out)

def to_excel_bytes(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "분석 결과"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # 보기 좋게 폭 조정
    for col_idx, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(28, len(col)+4))
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
